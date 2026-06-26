"""Bulk CSV import endpoints (post-Phase-6).

For each kind (sites, trials, projections):
  * happy-path: preview reports actions, commit writes them
  * at least one targeted error case (unknown FK, target-sum mismatch,
    non-Monday week, duplicate row, ...)
  * commit transaction rolls back when any row fails (no partial state)
"""

from __future__ import annotations

from datetime import date, timedelta
from io import BytesIO

import pytest
from httpx import AsyncClient
from openpyxl import Workbook, load_workbook

from app.models.soa_parse_job import SoaParseJob  # noqa: F401 — register table


def _next_monday() -> date:
    today = date.today()
    return today + timedelta(days=(7 - today.weekday()) % 7 or 7)


async def _signup(client: AsyncClient, name: str) -> dict:
    res = await client.post(
        "/orgs",
        json={
            "org_name": name,
            "default_timezone": "America/New_York",
            "admin_email": f"admin@{name.lower()}.example.com",
            "admin_password": "correct-horse-battery-staple",
            "admin_name": f"{name} admin",
        },
    )
    assert res.status_code == 201
    return res.json()


async def _login(client: AsyncClient, org_id: str, email: str) -> None:
    res = await client.post(
        "/auth/login",
        json={"email": email, "password": "correct-horse-battery-staple", "org_id": org_id},
    )
    assert res.status_code == 204


def _upload(name: str, body: str) -> dict:
    return {"file": (name, BytesIO(body.encode("utf-8")), "text/csv")}


@pytest.fixture
async def org(client: AsyncClient) -> dict:
    o = await _signup(client, "ImpOrg")
    await _login(client, o["id"], "admin@imporg.example.com")
    return o


# ---------------------------------------------------------------------------
# Sites


async def test_sites_template_download(client: AsyncClient, org: dict) -> None:
    res = await client.get("/imports/templates/sites.csv")
    assert res.status_code == 200
    assert "name,timezone" in res.text
    assert "sites-template.csv" in res.headers["content-disposition"]


async def test_sites_preview_and_commit_happy_path(
    client: AsyncClient, org: dict
) -> None:
    csv = (
        "name,timezone,operating_weekdays,hours_per_day,rooms\n"
        "Alpha Site,America/New_York,Mon Tue Wed Thu Fri,10,2\n"
        'Beta Site,America/Chicago,"0,1,2,3,4",9,1\n'
    )
    pv = await client.post("/imports/sites/preview", files=_upload("sites.csv", csv))
    assert pv.status_code == 200, pv.text
    body = pv.json()
    assert body["ok"] is True
    assert len(body["actions"]) == 2
    assert body["errors"] == []

    # Preview is read-only — sites listing still empty.
    assert (await client.get("/sites")).json() == []

    cm = await client.post("/imports/sites/commit", files=_upload("sites.csv", csv))
    assert cm.status_code == 200, cm.text
    names = sorted(s["name"] for s in (await client.get("/sites")).json())
    assert names == ["Alpha Site", "Beta Site"]


async def test_sites_commit_rolls_back_on_any_error(
    client: AsyncClient, org: dict
) -> None:
    csv = (
        "name,timezone,operating_weekdays,hours_per_day,rooms\n"
        "Good Site,America/New_York,Mon Tue Wed Thu Fri,10,2\n"
        "Bad Site,America/New_York,Mon Tue Wed Thu Fri,99,2\n"  # hours out of range
    )
    res = await client.post("/imports/sites/commit", files=_upload("s.csv", csv))
    assert res.status_code == 422
    detail = res.json()["detail"]
    assert any("hours_per_day" in e["message"] for e in detail["errors"])
    # Good Site must NOT have been written — proves the transaction held.
    assert (await client.get("/sites")).json() == []


async def test_sites_duplicate_in_file_blocked(
    client: AsyncClient, org: dict
) -> None:
    csv = (
        "name,timezone,operating_weekdays,hours_per_day,rooms\n"
        "Dup,America/New_York,Mon Tue Wed Thu Fri,10,2\n"
        "Dup,America/New_York,Mon Tue Wed Thu Fri,10,2\n"
    )
    pv = await client.post("/imports/sites/preview", files=_upload("d.csv", csv))
    assert pv.status_code == 200
    assert any("more than once" in e["message"] for e in pv.json()["errors"])


# ---------------------------------------------------------------------------
# Trials


async def _seed_two_sites_and_a_curve(client: AsyncClient) -> None:
    for n in ("S-A", "S-B"):
        await client.post(
            "/sites",
            json={
                "name": n,
                "timezone": "America/New_York",
                "operating_weekdays": [0, 1, 2, 3, 4],
                "hours_per_day": 10,
                "rooms": 2,
            },
        )
    await client.post(
        "/attrition-curves", json={"name": "Custom", "total_dropout_pct": 0.1}
    )


async def test_trials_happy_path_creates_draft_with_arm_and_assignments(
    client: AsyncClient, org: dict
) -> None:
    await _seed_two_sites_and_a_curve(client)
    csv = (
        "name,sponsor,fpfv,lpfv,lplv,enrollment_target,screening_target,"
        "attrition_curve_name,site_name,per_site_enrollment_target,"
        "per_site_screening_target\n"
        "T1,Acme,2026-09-07,2027-09-06,2028-09-04,200,250,Custom,S-A,100,125\n"
        "T1,,,,,,,,S-B,100,125\n"
    )
    pv = await client.post("/imports/trials/preview", files=_upload("t.csv", csv))
    assert pv.status_code == 200, pv.text
    pv_body = pv.json()
    assert pv_body["ok"] is True
    assert "2 site assignment" in pv_body["actions"][0]

    cm = await client.post("/imports/trials/commit", files=_upload("t.csv", csv))
    assert cm.status_code == 200, cm.text
    trials = (await client.get("/trials")).json()
    assert len(trials) == 1
    t = trials[0]
    assert t["name"] == "T1"
    assert t["status"] == "draft"  # PRD §6.2 — activation stays with the user
    # Default arm was auto-created (matches /trials POST behavior).
    arms = (await client.get(f"/trials/{t['id']}/arms")).json()
    assert [a["name"] for a in arms] == ["Default Arm"]
    # Both sites assigned.
    sts = (await client.get(f"/trials/{t['id']}/sites")).json()
    assert sorted(a["per_site_enrollment_target"] for a in sts) == [100, 100]


async def test_trials_target_sum_mismatch_blocked(
    client: AsyncClient, org: dict
) -> None:
    await _seed_two_sites_and_a_curve(client)
    csv = (
        "name,sponsor,fpfv,lpfv,lplv,enrollment_target,screening_target,"
        "attrition_curve_name,site_name,per_site_enrollment_target,"
        "per_site_screening_target\n"
        # study says 200/250 but sites sum to 150/180 — should error
        "T2,Acme,2026-09-07,2027-09-06,2028-09-04,200,250,Custom,S-A,80,90\n"
        "T2,,,,,,,,S-B,70,90\n"
    )
    pv = await client.post("/imports/trials/preview", files=_upload("t.csv", csv))
    msgs = " ".join(e["message"] for e in pv.json()["errors"])
    assert "rand sum 150" in msgs
    assert "screen sum 180" in msgs


async def test_trials_continuation_row_with_conflicting_fields_blocked(
    client: AsyncClient, org: dict
) -> None:
    await _seed_two_sites_and_a_curve(client)
    csv = (
        "name,sponsor,fpfv,lpfv,lplv,enrollment_target,screening_target,"
        "attrition_curve_name,site_name,per_site_enrollment_target,"
        "per_site_screening_target\n"
        "T3,Acme,2026-09-07,2027-09-06,2028-09-04,200,250,Custom,S-A,100,125\n"
        # Sponsor disagrees with first row's "Acme" — should error
        "T3,DifferentSponsor,,,,,,,S-B,100,125\n"
    )
    pv = await client.post("/imports/trials/preview", files=_upload("t.csv", csv))
    assert any("sponsor differs" in e["message"] for e in pv.json()["errors"])


async def test_trials_unknown_site_blocked(client: AsyncClient, org: dict) -> None:
    await _seed_two_sites_and_a_curve(client)
    csv = (
        "name,sponsor,fpfv,lpfv,lplv,enrollment_target,screening_target,"
        "attrition_curve_name,site_name,per_site_enrollment_target,"
        "per_site_screening_target\n"
        "T4,Acme,2026-09-07,2027-09-06,2028-09-04,100,125,Custom,Nowhere,100,125\n"
    )
    res = await client.post("/imports/trials/commit", files=_upload("t.csv", csv))
    assert res.status_code == 422
    assert any(
        "unknown site" in e["message"] for e in res.json()["detail"]["errors"]
    )
    # And nothing was written.
    assert (await client.get("/trials")).json() == []


# ---------------------------------------------------------------------------
# Projections


async def _seed_trial_with_arm(
    client: AsyncClient,
) -> tuple[str, str, str]:
    """Returns (site_id, trial_id, arm_id) for a draft trial with one default arm."""
    await client.post(
        "/sites",
        json={
            "name": "PSite",
            "timezone": "America/New_York",
            "operating_weekdays": [0, 1, 2, 3, 4],
            "hours_per_day": 10,
            "rooms": 1,
        },
    )
    site_id = (await client.get("/sites")).json()[0]["id"]
    curve = (
        await client.post(
            "/attrition-curves", json={"name": "Zero", "total_dropout_pct": 0.0}
        )
    ).json()
    trial = (
        await client.post(
            "/trials",
            json={
                "name": "PT",
                "fpfv": "2026-09-07",
                "lpfv": "2027-09-06",
                "lplv": "2028-09-04",
                "enrollment_target": 50,
                "screening_target": 60,
                "attrition_curve_id": curve["id"],
            },
        )
    ).json()
    arm = (await client.get(f"/trials/{trial['id']}/arms")).json()[0]
    await client.post(
        f"/trials/{trial['id']}/sites",
        json={
            "site_id": site_id,
            "per_site_enrollment_target": 50,
            "per_site_screening_target": 60,
        },
    )
    return site_id, trial["id"], arm["id"]


async def test_projections_happy_path(client: AsyncClient, org: dict) -> None:
    await _seed_trial_with_arm(client)
    mon = _next_monday()
    csv = (
        "site_name,trial_name,arm_name,week_start,proj_screened,proj_randomized\n"
        f"PSite,PT,,{mon.isoformat()},5,4\n"
        f"PSite,PT,,{(mon + timedelta(days=7)).isoformat()},6,5\n"
    )
    pv = await client.post(
        "/imports/projections/preview", files=_upload("p.csv", csv)
    )
    assert pv.status_code == 200
    assert pv.json()["ok"] is True

    cm = await client.post(
        "/imports/projections/commit", files=_upload("p.csv", csv)
    )
    assert cm.status_code == 200, cm.text


async def test_projections_non_monday_blocked(
    client: AsyncClient, org: dict
) -> None:
    await _seed_trial_with_arm(client)
    mon = _next_monday()
    tue = (mon + timedelta(days=1)).isoformat()
    csv = (
        "site_name,trial_name,arm_name,week_start,proj_screened,proj_randomized\n"
        f"PSite,PT,,{tue},5,4\n"
    )
    res = await client.post(
        "/imports/projections/commit", files=_upload("p.csv", csv)
    )
    assert res.status_code == 422
    assert any(
        "must be a Monday" in e["message"]
        for e in res.json()["detail"]["errors"]
    )


async def test_projections_upsert_overwrites_existing_week(
    client: AsyncClient, org: dict
) -> None:
    """Re-importing the same (site, trial, arm, week) updates the projection.
    This is the load-bearing UX: an operator fixes a typo in the spreadsheet
    and re-uploads. Actuals (set elsewhere) must NOT be clobbered."""
    site_id, trial_id, arm_id = await _seed_trial_with_arm(client)
    mon = _next_monday()
    csv1 = (
        "site_name,trial_name,arm_name,week_start,proj_screened,proj_randomized\n"
        f"PSite,PT,,{mon.isoformat()},5,4\n"
    )
    await client.post("/imports/projections/commit", files=_upload("p.csv", csv1))

    csv2 = (
        "site_name,trial_name,arm_name,week_start,proj_screened,proj_randomized\n"
        f"PSite,PT,,{mon.isoformat()},9,8\n"
    )
    res = await client.post(
        "/imports/projections/commit", files=_upload("p2.csv", csv2)
    )
    assert res.status_code == 200, res.text

    # Read back via the existing enrollment-weeks endpoint.
    sts = (await client.get(f"/trials/{trial_id}/sites")).json()
    site_trial_id = sts[0]["id"]
    weeks = (
        await client.get(
            f"/site-trials/{site_trial_id}/enrollment-weeks",
            params={
                "arm_id": arm_id,
                "from": mon.isoformat(),
                "to": mon.isoformat(),
            },
        )
    ).json()
    assert len(weeks) == 1
    assert weeks[0]["proj_screened"] == 9
    assert weeks[0]["proj_randomized"] == 8


# ---------------------------------------------------------------------------
# Auth + RLS


async def test_xlsx_template_trials_includes_site_reference_sheet(
    client: AsyncClient, org: dict
) -> None:
    """The trials template's Reference sheet lists existing site + curve
    names — this is the user's defense against typos like a trailing space."""
    await _seed_two_sites_and_a_curve(client)
    res = await client.get("/imports/templates/trials.xlsx")
    assert res.status_code == 200
    assert "spreadsheetml" in res.headers["content-type"]
    wb = load_workbook(BytesIO(res.content))
    assert wb.sheetnames == ["Template", "Reference"]

    ref = wb["Reference"]
    rows = [list(r) for r in ref.iter_rows(values_only=True)]
    # Header row + one data row per site (we seeded 2, sorted).
    assert rows[0] == ["Existing site names", "Existing attrition curve names"]
    site_col = [r[0] for r in rows[1:]]
    curve_col = [r[1] for r in rows[1:]]
    assert site_col[:2] == ["S-A", "S-B"]
    assert "Custom" in curve_col


async def test_xlsx_template_projections_is_per_site_grid(
    client: AsyncClient, org: dict
) -> None:
    """The projections template is one sheet per site, app-shaped: each
    assigned study is a Screened/Randomized column group with weeks down the
    rows (PRD §7.3)."""
    await _seed_trial_with_arm(client)
    res = await client.get("/imports/templates/projections.xlsx")
    assert res.status_code == 200
    wb = load_workbook(BytesIO(res.content))

    assert "Instructions" in wb.sheetnames
    # The site gets its own tab; A1 carries the full site name.
    site_ws = next(ws for ws in wb.worksheets if ws["A1"].value == "PSite")
    assert site_ws["B1"].value == "PT"  # study (trial) group header
    assert site_ws["B2"].value == "Default Arm"  # arm header
    assert site_ws["A3"].value == "week_start"
    assert site_ws["B3"].value == "Screened"
    assert site_ws["C3"].value == "Randomized"
    # 52 Monday rows, all Mondays, starting row 4.
    week_cells = [site_ws.cell(row=r, column=1).value for r in range(4, 56)]
    assert len(week_cells) == 52
    assert all(date.fromisoformat(str(w)).weekday() == 0 for w in week_cells)

    # The study's cells (Screened/Randomized) are highlighted for weeks inside
    # its FPFV–LPFV window (seed: 2026-09-07 .. 2027-09-06); weeks before FPFV
    # are not. Column A (week_start) is left plain.
    fpfv = date(2026, 9, 7)
    rows = {
        r: date.fromisoformat(str(site_ws.cell(row=r, column=1).value))
        for r in range(4, 56)
    }
    before = next(r for r, d in rows.items() if d < fpfv)
    inside = next(r for r, d in rows.items() if d >= fpfv)
    # PT's Screened (col B) + Randomized (col C) highlighted only inside the window.
    assert site_ws.cell(row=before, column=2).fill.patternType is None
    assert site_ws.cell(row=inside, column=2).fill.patternType == "solid"
    assert site_ws.cell(row=inside, column=3).fill.patternType == "solid"
    # week_start column is no longer emphasized.
    assert site_ws.cell(row=inside, column=1).fill.patternType is None
    assert not site_ws.cell(row=inside, column=1).font.bold


async def test_xlsx_projections_grid_round_trips_to_enrollment_weeks(
    client: AsyncClient, org: dict
) -> None:
    """Download the per-site grid, fill two weeks for the assigned study,
    upload, and the projections land as EnrollmentWeek rows."""
    site_id, trial_id, arm_id = await _seed_trial_with_arm(client)
    res = await client.get("/imports/templates/projections.xlsx")
    wb = load_workbook(BytesIO(res.content))
    ws = next(w for w in wb.worksheets if w["A1"].value == "PSite")

    # Fill the first two week rows: Screened=B, Randomized=C.
    w0 = date.fromisoformat(str(ws["A4"].value))
    w1 = date.fromisoformat(str(ws["A5"].value))
    ws["B4"], ws["C4"] = 7, 5
    ws["B5"], ws["C5"] = 8, 6
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    files = {
        "file": (
            "projections-filled.xlsx",
            buf,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    cm = await client.post("/imports/projections/commit", files=files)
    assert cm.status_code == 200, cm.text

    # Read the enrollment weeks back for this (site_trial, arm).
    st = (await client.get(f"/trials/{trial_id}/sites")).json()[0]
    weeks = (
        await client.get(
            f"/site-trials/{st['id']}/enrollment-weeks",
            params={
                "arm_id": arm_id,
                "from": w0.isoformat(),
                "to": w1.isoformat(),
            },
        )
    ).json()
    by_week = {w["week_start"]: w for w in weeks}
    assert by_week[w0.isoformat()]["proj_screened"] == 7
    assert by_week[w0.isoformat()]["proj_randomized"] == 5
    assert by_week[w1.isoformat()]["proj_screened"] == 8
    assert by_week[w1.isoformat()]["proj_randomized"] == 6


async def test_xlsx_upload_round_trips_through_validator(
    client: AsyncClient, org: dict
) -> None:
    """The user can fill in the downloaded XLSX and upload it directly —
    the server normalizes XLSX → CSV before validation runs."""
    # Build a small in-memory XLSX equivalent to a clean Sites CSV.
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Template"
    ws.append(
        ["name", "timezone", "operating_weekdays", "hours_per_day", "rooms"]
    )
    ws.append(["Alpha XLSX", "America/New_York", "Mon Tue Wed Thu Fri", 10, 2])
    ws.append(["Beta XLSX", "America/Chicago", "0,1,2,3,4", 9, 1])
    # Reference sheet — should be IGNORED on upload.
    ref = wb.create_sheet("Reference")
    ref.append(["should not", "be parsed"])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    files = {
        "file": (
            "sites-filled.xlsx",
            buf,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    pv = await client.post("/imports/sites/preview", files=files)
    assert pv.status_code == 200, pv.text
    body = pv.json()
    assert body["ok"] is True
    assert len(body["actions"]) == 2

    buf.seek(0)
    cm = await client.post("/imports/sites/commit", files=files)
    assert cm.status_code == 200, cm.text
    names = sorted(s["name"] for s in (await client.get("/sites")).json())
    assert names == ["Alpha XLSX", "Beta XLSX"]


async def test_preview_requires_admin(client: AsyncClient, org: dict) -> None:
    # Promote a second user, then demote our admin away from org_admin.
    other = await client.post(
        "/users",
        json={
            "email": "second@imporg.example.com",
            "name": "Second",
            "password": "correct-horse-battery-staple",
            "role": "org_admin",
        },
    )
    assert other.status_code == 201
    me = (await client.get("/auth/me")).json()
    await client.patch(f"/users/{me['user_id']}", json={"role": "viewer"})

    res = await client.post(
        "/imports/sites/preview",
        files=_upload("s.csv", "name,timezone,operating_weekdays,hours_per_day,rooms\n"),
    )
    assert res.status_code == 403
