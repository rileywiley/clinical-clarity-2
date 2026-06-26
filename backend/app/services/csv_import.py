"""Bulk CSV import — sites, trials, projections.

Post-Phase-6 commercialization feature. Three kinds, each with a
preview + commit pair. Validation is all-or-nothing per upload: if any
row fails, the whole file is rejected and no writes happen (PRD §6.2
modeling decisions stay intact).

Foreign keys resolve by name within the org. Unknown names → preview
error. Trials always import as ``draft``: bulk import seeds the trial,
its site assignments, and an attrition curve, but never an SoA — and an
SoA (with a randomization visit) is required to become ``planned`` or
``active``. So the validated transition (`/trials/{id}/plan` or
`/activate`, PRD §6.9 / §7.1) still owns the lifecycle; a draft can't be
shortcut into the forecast by import. Projections, by contrast, can be
imported against a trial in any status (planned trials carry their
future-dated projections too — PRD §7.1).
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from typing import Literal
from uuid import UUID

from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.attrition_curve import AttritionCurve
from app.models.enrollment_week import EnrollmentWeek
from app.models.org_settings import OrgSettings
from app.models.site import Site
from app.models.site_trial import SiteTrial
from app.models.trial import Arm, Trial, TrialStatus

ImportKind = Literal["sites", "trials", "projections"]


# ---------------------------------------------------------------------------
# Shared types


@dataclass
class ImportError:
    """One human-readable validation problem, anchored to a CSV row."""

    row: int  # 1-indexed; row 1 is the header
    message: str


@dataclass
class ImportPreview:
    actions: list[str] = field(default_factory=list)
    errors: list[ImportError] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return not self.errors


# ---------------------------------------------------------------------------
# Templates — header + a couple of example rows


_TEMPLATES: dict[ImportKind, str] = {
    "sites": (
        "name,timezone,operating_weekdays,hours_per_day,rooms\n"
        "NYU Langone,America/New_York,Mon Tue Wed Thu Fri,10,3\n"
        "Mayo Rochester,America/Chicago,\"Mon,Tue,Wed,Thu,Fri\",9,2\n"
    ),
    "trials": (
        "name,sponsor,fpfv,lpfv,lplv,enrollment_target,screening_target,"
        "attrition_curve_name,site_name,per_site_enrollment_target,"
        "per_site_screening_target\n"
        "TRIAL-42,Acme Pharma,2026-08-03,2027-08-02,2028-08-07,200,250,"
        "Standard,NYU Langone,100,125\n"
        "TRIAL-42,,,,,,,,Mayo Rochester,100,125\n"
    ),
    "projections": (
        "site_name,trial_name,arm_name,week_start,proj_screened,proj_randomized\n"
        "NYU Langone,TRIAL-42,,2026-08-03,5,4\n"
        "NYU Langone,TRIAL-42,,2026-08-10,5,4\n"
    ),
}


def template_for(kind: ImportKind) -> str:
    return _TEMPLATES[kind]


# ---------------------------------------------------------------------------
# XLSX templates — same data as the CSV templates, but with a second
# **Reference** sheet that lists the existing org-scoped names a user
# needs to type into the template (sites for trials, trials for
# projections). This is what stops the most common upload error
# ("unknown site 'NYU Langone '" because of a trailing space or a typo).


def _set_header_style(ws) -> None:
    from openpyxl.styles import Font, PatternFill

    fill = PatternFill("solid", fgColor="E2E8F0")  # slate-200
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = bold


def _autosize(ws) -> None:
    from openpyxl.utils import get_column_letter

    # Index-based column letter: merged cells yield MergedCell objects with no
    # `.column_letter`, so we can't read it off the first cell of each column.
    for i, col in enumerate(ws.columns, start=1):
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(i)].width = min(max(12, max_len + 2), 40)


async def template_xlsx_for(
    kind: ImportKind, org_id: UUID, db: AsyncSession
) -> bytes:
    """Build the kind's XLSX template, populating the Reference sheet with
    live names from the org. Returns the binary .xlsx bytes."""
    # Projections get a richer, app-shaped workbook: one sheet per site, with
    # each assigned study as a Screened/Randomized column group and weeks down
    # the rows (mirrors the Projections page, PRD §7.3). Handled separately.
    if kind == "projections":
        return await _projections_grid_template(org_id, db)

    wb = Workbook()
    data_ws = wb.active
    assert data_ws is not None
    data_ws.title = "Template"

    # Mirror the CSV template — header + example rows.
    reader = csv.reader(io.StringIO(_TEMPLATES[kind]))
    for row in reader:
        data_ws.append(row)
    _set_header_style(data_ws)
    _autosize(data_ws)

    # Reference sheet — what the user needs to type into the data sheet.
    ref_ws = wb.create_sheet("Reference")
    if kind == "trials":
        sites = (
            await db.execute(
                select(Site.name).where(Site.org_id == org_id).order_by(Site.name)
            )
        ).scalars().all()
        curves = (
            await db.execute(
                select(AttritionCurve.name)
                .where(AttritionCurve.org_id == org_id)
                .order_by(AttritionCurve.name)
            )
        ).scalars().all()
        ref_ws.append(["Existing site names", "Existing attrition curve names"])
        for i in range(max(len(sites), len(curves))):
            ref_ws.append(
                [
                    sites[i] if i < len(sites) else "",
                    curves[i] if i < len(curves) else "",
                ]
            )
        if not sites and not curves:
            ref_ws.append(["(no sites yet — create one first)", ""])
    else:
        # Sites template needs no reference data — nothing exists to
        # collide with. Show a small note instead so users aren't
        # confused by an empty sheet.
        ref_ws.append(
            ["The Sites template needs no reference — site names are unique within an org."]
        )

    _set_header_style(ref_ws)
    _autosize(ref_ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Projections grid template — one sheet per site, app-shaped (PRD §7.3).
#
# Layout per site sheet (a 3-row header so trial + arm survive round-trip):
#   A1 = <site name>      B1:C1 = <trial>     D1:E1 = <trial> ...
#   A2 = (blank)          B2:C2 = <arm>       D2:E2 = <arm>   ...
#   A3 = "week_start"     B3 = "Screened" C3 = "Randomized"   ...
#   A4.. = Monday dates; the Screened/Randomized cells are left blank to fill.
# Each assigned study (one column group per trial×arm) is listed across the top
# so the sheet reads like the Projections page. Re-importing upserts.

_ILLEGAL_SHEET_CHARS = set(r"[]:*?/\\")
_PROJECTION_TEMPLATE_WEEKS = 52


def _safe_sheet_title(name: str, used: set[str]) -> str:
    """Excel sheet titles are ≤31 chars, can't contain []:*?/\\, and must be
    unique. The full site name still lives in cell A1, so a lossy title is fine."""
    cleaned = "".join(c for c in name if c not in _ILLEGAL_SHEET_CHARS).strip() or "Site"
    title = cleaned[:31]
    i = 2
    while title.lower() in used:
        suffix = f" ({i})"
        title = cleaned[: 31 - len(suffix)] + suffix
        i += 1
    used.add(title.lower())
    return title


def _template_mondays(today: date | None = None) -> list[date]:
    """Current Monday + the next 51 (a full year of forward weeks)."""
    base = today or date.today()
    monday = base - timedelta(days=base.weekday())
    return [monday + timedelta(weeks=i) for i in range(_PROJECTION_TEMPLATE_WEEKS)]


async def _studies_by_site(
    org_id: UUID, db: AsyncSession
) -> dict[UUID, list[tuple[Trial, Arm]]]:
    """For each site, the (trial, arm) column groups it should show — every
    active assignment, each trial expanded to its arms. Ordered by trial then
    arm name for a stable layout."""
    assignments = (
        await db.execute(
            select(SiteTrial.site_id, Trial)
            .join(Trial, Trial.id == SiteTrial.trial_id)
            .where(SiteTrial.org_id == org_id, SiteTrial.active.is_(True))
        )
    ).all()
    arms_by_trial: dict[UUID, list[Arm]] = {}
    for arm in (await db.execute(select(Arm).where(Arm.org_id == org_id))).scalars().all():
        arms_by_trial.setdefault(arm.trial_id, []).append(arm)

    out: dict[UUID, list[tuple[Trial, Arm]]] = {}
    for site_id, trial in assignments:
        for arm in arms_by_trial.get(trial.id, []):
            out.setdefault(site_id, []).append((trial, arm))
    for groups in out.values():
        groups.sort(key=lambda ta: (ta[0].name.lower(), ta[1].name.lower()))
    return out


async def _projections_grid_template(org_id: UUID, db: AsyncSession) -> bytes:
    from openpyxl.styles import Font, PatternFill

    fill = PatternFill("solid", fgColor="E2E8F0")  # slate-200 header fill
    window_fill = PatternFill("solid", fgColor="FEF3C7")  # amber-100 — in-window cells
    bold = Font(bold=True)

    wb = Workbook()
    instr = wb.active
    assert instr is not None
    instr.title = "Instructions"
    for line in (
        ["Projections import — one tab per site"],
        [],
        ["• Each site has its own tab. Don't rename tabs or move the three header rows."],
        ["• Row 1 = study, Row 2 = arm, Row 3 = column labels."],
        ["• Enter weekly projected counts under each study's Screened / Randomized columns."],
        ["• week_start dates are Mondays (52 forward weeks). Leave a cell blank for 'no projection' that week."],
        ["• Highlighted cells fall inside a study's enrollment window (FPFV–LPFV) — that's where projections belong."],
        ["• Re-importing overwrites the same (site, study, arm, week). Sites/studies you"],
        ["  don't touch are untouched."],
    ):
        instr.append(line)
    instr["A1"].font = bold

    sites = (
        await db.execute(select(Site).where(Site.org_id == org_id).order_by(Site.name))
    ).scalars().all()
    groups_by_site = await _studies_by_site(org_id, db)
    mondays = _template_mondays()
    used_titles = {"instructions"}

    if not sites:
        ws = wb.create_sheet("No sites")
        ws.append(["No sites yet — create a site (and assign studies) first."])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    for site in sites:
        ws = wb.create_sheet(_safe_sheet_title(site.name, used_titles))
        groups = groups_by_site.get(site.id, [])

        row1: list[object] = [site.name]
        row2: list[object] = [""]
        row3: list[object] = ["week_start"]
        for trial, arm in groups:
            row1 += [trial.name, ""]
            row2 += [arm.name, ""]
            row3 += ["Screened", "Randomized"]
        ws.append(row1)
        ws.append(row2)
        ws.append(row3)
        for m in mondays:
            ws.append([m.isoformat()] + [""] * (2 * len(groups)))

        # Merge + style each study's 2-col group header; bold the header rows.
        for gi in range(len(groups)):
            c = 2 + 2 * gi
            ws.merge_cells(start_row=1, start_column=c, end_row=1, end_column=c + 1)
            ws.merge_cells(start_row=2, start_column=c, end_row=2, end_column=c + 1)
        for r in (1, 2, 3):
            for cell in ws[r]:
                cell.font = bold
                cell.fill = fill

        # Highlight the cells under each study for the weeks inside that study's
        # enrollment window (FPFV–LPFV) — that's where projections belong. The
        # fill lands on the study's own Screened/Randomized input cells, so the
        # emphasis sits exactly where the user types.
        for ri, m in enumerate(mondays):
            r = 4 + ri
            for gi, (trial, _arm) in enumerate(groups):
                if trial.fpfv <= m <= trial.lpfv:
                    sc = 2 + 2 * gi
                    ws.cell(row=r, column=sc).fill = window_fill
                    ws.cell(row=r, column=sc + 1).fill = window_fill

        ws.freeze_panes = "B4"  # keep week_start col + header rows visible
        _autosize(ws)

        if not groups:
            ws.append([])
            ws.append(["(no studies assigned to this site yet)"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Upload normalization — let users hand us either CSV bytes or the
# XLSX template back (whether or not they edited the Reference sheet).
# Everything downstream operates on a CSV string.


def normalize_upload(filename: str, raw: bytes) -> str:
    """Return CSV text for the validator. Detects XLSX by extension.

    Two XLSX shapes are accepted:
      - **Flat** (sites/trials, and the legacy projections CSV): a single data
        sheet whose first row is the column header. We flatten the first sheet.
      - **Projections grid** (the new per-site template): one sheet per site
        with study column groups. We flatten *all* site sheets into the same
        legacy ``site_name,trial_name,arm_name,week_start,...`` CSV the
        projections validator already understands, so nothing downstream changes.
    The Reference / Instructions sheet is ignored on upload either way."""
    is_xlsx = filename.lower().endswith(".xlsx")
    # openpyxl reads .xlsx; passing CSV bytes through would crash, so we
    # branch on extension. (Sniffing magic bytes is brittle.)
    if not is_xlsx:
        return raw.decode("utf-8-sig", errors="replace")
    # Not read_only: the grid flattener needs random cell access, and these
    # workbooks are tiny (per-site sheets, ~12 rows).
    wb = load_workbook(io.BytesIO(raw), data_only=True)
    if _looks_like_projection_grid(wb):
        return _flatten_projection_grid(wb)
    ws = wb.worksheets[0]
    out = io.StringIO()
    writer = csv.writer(out)
    for row in ws.iter_rows(values_only=True):
        if all(v is None or v == "" for v in row):
            continue
        writer.writerow(["" if v is None else _xlsx_cell_to_str(v) for v in row])
    return out.getvalue()


def _looks_like_projection_grid(wb) -> bool:
    """A per-site projections grid sheet carries the literal ``week_start`` in
    cell A3 (row 3). The flat templates never do (their headers are in row 1),
    so this cleanly distinguishes the two without sniffing."""
    for ws in wb.worksheets:
        if ws.title.strip().lower() == "instructions":
            continue
        if str(ws.cell(row=3, column=1).value or "").strip().lower() == "week_start":
            return True
    return False


def _flatten_projection_grid(wb) -> str:
    """Flatten the per-site grid workbook into the legacy projections CSV.

    For each site sheet: site name from A1, study from row 1, arm from row 2,
    Screened/Randomized from the column pair under each study, weeks from
    column A (row 4 down). A (week, study) with *both* cells blank is skipped
    (no projection); a half-filled pair fills the blank side with 0."""
    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow(
        ["site_name", "trial_name", "arm_name", "week_start",
         "proj_screened", "proj_randomized"]
    )
    for ws in wb.worksheets:
        if ws.title.strip().lower() == "instructions":
            continue
        if str(ws.cell(row=3, column=1).value or "").strip().lower() != "week_start":
            continue
        site_name = str(ws.cell(row=1, column=1).value or "").strip()
        if not site_name:
            continue

        # Discover study column groups: row 1 trial name at col 2, 4, 6, …
        groups: list[tuple[str, str, int]] = []  # (trial, arm, screened_col)
        col = 2
        while True:
            trial = ws.cell(row=1, column=col).value
            if trial is None or str(trial).strip() == "":
                break
            arm = ws.cell(row=2, column=col).value
            groups.append(
                (str(trial).strip(), str(arm).strip() if arm else "Default Arm", col)
            )
            col += 2
        if not groups:
            continue

        for r in range(4, ws.max_row + 1):
            wk = ws.cell(row=r, column=1).value
            if wk is None or str(wk).strip() == "":
                continue
            wk_str = _xlsx_cell_to_str(wk)
            for trial, arm, sc in groups:
                sv = ws.cell(row=r, column=sc).value
                rv = ws.cell(row=r, column=sc + 1).value
                s_blank = sv is None or str(sv).strip() == ""
                r_blank = rv is None or str(rv).strip() == ""
                if s_blank and r_blank:
                    continue
                writer.writerow(
                    [site_name, trial, arm, wk_str,
                     "0" if s_blank else _xlsx_num_to_str(sv),
                     "0" if r_blank else _xlsx_num_to_str(rv)]
                )
    return out.getvalue()


def _xlsx_cell_to_str(v: object) -> str:
    """Excel dates round-trip as datetime objects; we need ISO strings
    back so the existing date validator stays happy."""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return str(v)


def _xlsx_num_to_str(v: object) -> str:
    """Numeric projection cells come back as int or float (Excel stores all
    numbers as doubles). Render whole numbers without a trailing '.0' so the
    integer validator stays happy; leave non-numerics for it to reject."""
    if isinstance(v, bool):  # bool is an int subclass — guard first
        return str(v)
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


# ---------------------------------------------------------------------------
# Sites


_DAY_TOKENS: dict[str, int] = {
    "mon": 0, "tue": 1, "wed": 2, "thu": 3, "fri": 4, "sat": 5, "sun": 6,
    "monday": 0, "tuesday": 1, "wednesday": 2, "thursday": 3,
    "friday": 4, "saturday": 5, "sunday": 6,
    "0": 0, "1": 1, "2": 2, "3": 3, "4": 4, "5": 5, "6": 6,
}


def _parse_weekdays(raw: str, row: int, errors: list[ImportError]) -> list[int]:
    tokens = [t.strip().lower() for t in raw.replace(",", " ").split() if t.strip()]
    if not tokens:
        errors.append(ImportError(row, "operating_weekdays cannot be empty"))
        return []
    out: list[int] = []
    for tok in tokens:
        if tok not in _DAY_TOKENS:
            errors.append(
                ImportError(row, f"operating_weekdays: unknown token {tok!r}")
            )
            continue
        d = _DAY_TOKENS[tok]
        if d not in out:
            out.append(d)
    return sorted(out)


def _parse_int(raw: str, field_name: str, row: int, errors: list[ImportError]) -> int | None:
    try:
        return int(raw.strip())
    except (ValueError, AttributeError):
        errors.append(ImportError(row, f"{field_name}: not an integer ({raw!r})"))
        return None


def _parse_date(raw: str, field_name: str, row: int, errors: list[ImportError]) -> date | None:
    try:
        return datetime.strptime(raw.strip(), "%Y-%m-%d").date()
    except (ValueError, AttributeError):
        errors.append(
            ImportError(row, f"{field_name}: expected YYYY-MM-DD ({raw!r})")
        )
        return None


@dataclass
class _SiteRow:
    name: str
    timezone: str
    operating_weekdays: list[int]
    hours_per_day: int
    rooms: int


async def _validate_sites(
    text: str, org_id: UUID, db: AsyncSession
) -> tuple[list[_SiteRow], list[ImportError]]:
    rows: list[_SiteRow] = []
    errors: list[ImportError] = []
    reader = csv.DictReader(io.StringIO(text))
    expected = {"name", "timezone", "operating_weekdays", "hours_per_day", "rooms"}
    if reader.fieldnames is None or set(reader.fieldnames) - expected:
        # Fieldnames may be a superset (Excel often appends blank columns);
        # only fail if the required ones are missing.
        missing = expected - set(reader.fieldnames or [])
        if missing:
            errors.append(
                ImportError(1, f"missing required columns: {sorted(missing)}")
            )
            return [], errors

    existing_names = set(
        (
            await db.execute(select(Site.name).where(Site.org_id == org_id))
        ).scalars().all()
    )
    seen_in_file: set[str] = set()
    for idx, raw in enumerate(reader, start=2):
        name = (raw.get("name") or "").strip()
        if not name:
            errors.append(ImportError(idx, "name is required"))
            continue
        if name in existing_names:
            errors.append(
                ImportError(idx, f"site {name!r} already exists in this org")
            )
            continue
        if name in seen_in_file:
            errors.append(
                ImportError(idx, f"site {name!r} appears more than once in the file")
            )
            continue
        seen_in_file.add(name)

        weekdays = _parse_weekdays(raw.get("operating_weekdays", ""), idx, errors)
        hours = _parse_int(raw.get("hours_per_day", ""), "hours_per_day", idx, errors)
        rooms = _parse_int(raw.get("rooms", ""), "rooms", idx, errors)
        tz = (raw.get("timezone") or "").strip()
        if not tz:
            errors.append(ImportError(idx, "timezone is required"))
            continue
        if hours is None or rooms is None or not weekdays:
            continue
        if not (1 <= hours <= 24):
            errors.append(ImportError(idx, "hours_per_day must be 1–24"))
            continue
        if rooms < 1:
            errors.append(ImportError(idx, "rooms must be ≥ 1"))
            continue
        rows.append(
            _SiteRow(
                name=name,
                timezone=tz,
                operating_weekdays=weekdays,
                hours_per_day=hours,
                rooms=rooms,
            )
        )
    return rows, errors


async def preview_sites(text: str, org_id: UUID, db: AsyncSession) -> ImportPreview:
    rows, errors = await _validate_sites(text, org_id, db)
    return ImportPreview(
        actions=[f"Create site {r.name!r} ({r.rooms} rooms × {r.hours_per_day}h)" for r in rows],
        errors=errors,
    )


async def commit_sites(text: str, org_id: UUID, db: AsyncSession) -> ImportPreview:
    rows, errors = await _validate_sites(text, org_id, db)
    if errors:
        return ImportPreview(actions=[], errors=errors)
    for r in rows:
        db.add(
            Site(
                org_id=org_id,
                name=r.name,
                timezone=r.timezone,
                operating_weekdays=r.operating_weekdays,
                hours_per_day=r.hours_per_day,
                rooms=r.rooms,
            )
        )
    await db.flush()
    return ImportPreview(actions=[f"Created site {r.name!r}" for r in rows], errors=[])


# ---------------------------------------------------------------------------
# Trials — multi-row format: one row per (trial, site) assignment.
# Trial-level fields may be omitted on rows 2+ for the same trial (inherit
# from the first occurrence). Sum of per-site targets must equal the
# trial's enrollment / screening target.


@dataclass
class _TrialBlock:
    name: str
    sponsor: str | None
    fpfv: date
    lpfv: date
    lplv: date
    enrollment_target: int
    screening_target: int
    attrition_curve_name: str
    assignments: list[tuple[str, int, int]]  # (site_name, rand, screen)


async def _validate_trials(
    text: str, org_id: UUID, db: AsyncSession
) -> tuple[list[_TrialBlock], list[ImportError]]:
    errors: list[ImportError] = []
    reader = csv.DictReader(io.StringIO(text))
    required = {
        "name", "sponsor", "fpfv", "lpfv", "lplv",
        "enrollment_target", "screening_target", "attrition_curve_name",
        "site_name", "per_site_enrollment_target", "per_site_screening_target",
    }
    missing = required - set(reader.fieldnames or [])
    if missing:
        errors.append(
            ImportError(1, f"missing required columns: {sorted(missing)}")
        )
        return [], errors

    existing_trials = set(
        (
            await db.execute(select(Trial.name).where(Trial.org_id == org_id))
        ).scalars().all()
    )
    sites_by_name = {
        s.name: s
        for s in (
            await db.execute(select(Site).where(Site.org_id == org_id))
        ).scalars().all()
    }
    curves_by_name = {
        c.name: c
        for c in (
            await db.execute(
                select(AttritionCurve).where(AttritionCurve.org_id == org_id)
            )
        ).scalars().all()
    }

    blocks: dict[str, _TrialBlock] = {}
    seen_site_per_trial: dict[str, set[str]] = {}

    for idx, raw in enumerate(reader, start=2):
        name = (raw.get("name") or "").strip()
        if not name:
            errors.append(ImportError(idx, "name is required on every row"))
            continue
        if name in existing_trials:
            errors.append(
                ImportError(idx, f"trial {name!r} already exists in this org")
            )
            continue

        site_name = (raw.get("site_name") or "").strip()
        if not site_name:
            errors.append(ImportError(idx, "site_name is required"))
            continue
        if site_name not in sites_by_name:
            errors.append(
                ImportError(idx, f"unknown site {site_name!r} (create it first)")
            )
            continue
        seen = seen_site_per_trial.setdefault(name, set())
        if site_name in seen:
            errors.append(
                ImportError(idx, f"site {site_name!r} appears twice for trial {name!r}")
            )
            continue
        seen.add(site_name)

        rand = _parse_int(
            raw.get("per_site_enrollment_target", ""),
            "per_site_enrollment_target",
            idx,
            errors,
        )
        screen = _parse_int(
            raw.get("per_site_screening_target", ""),
            "per_site_screening_target",
            idx,
            errors,
        )
        if rand is None or screen is None:
            continue
        if rand < 0 or screen < 0:
            errors.append(ImportError(idx, "per-site targets must be ≥ 0"))
            continue

        # First occurrence of this trial — basics are required here.
        if name not in blocks:
            sponsor = (raw.get("sponsor") or "").strip() or None
            fpfv = _parse_date(raw.get("fpfv", ""), "fpfv", idx, errors)
            lpfv = _parse_date(raw.get("lpfv", ""), "lpfv", idx, errors)
            lplv = _parse_date(raw.get("lplv", ""), "lplv", idx, errors)
            enr = _parse_int(
                raw.get("enrollment_target", ""), "enrollment_target", idx, errors
            )
            scr = _parse_int(
                raw.get("screening_target", ""), "screening_target", idx, errors
            )
            curve_name = (raw.get("attrition_curve_name") or "").strip()
            if not curve_name:
                errors.append(
                    ImportError(idx, "attrition_curve_name is required on first row")
                )
                continue
            if curve_name not in curves_by_name:
                errors.append(
                    ImportError(
                        idx,
                        f"unknown attrition curve {curve_name!r} (create it first)",
                    )
                )
                continue
            if not all([fpfv, lpfv, lplv]) or enr is None or scr is None:
                continue
            assert fpfv and lpfv and lplv
            if not (fpfv <= lpfv <= lplv):
                errors.append(
                    ImportError(idx, "fpfv ≤ lpfv ≤ lplv must hold")
                )
                continue
            blocks[name] = _TrialBlock(
                name=name,
                sponsor=sponsor,
                fpfv=fpfv,
                lpfv=lpfv,
                lplv=lplv,
                enrollment_target=enr,
                screening_target=scr,
                attrition_curve_name=curve_name,
                assignments=[(site_name, rand, screen)],
            )
        else:
            # Continuation row — trial-level fields may be blank (inherit)
            # or must equal the first row's value.
            b = blocks[name]
            for field_name, expected in (
                ("sponsor", b.sponsor or ""),
                ("fpfv", b.fpfv.isoformat()),
                ("lpfv", b.lpfv.isoformat()),
                ("lplv", b.lplv.isoformat()),
                ("enrollment_target", str(b.enrollment_target)),
                ("screening_target", str(b.screening_target)),
                ("attrition_curve_name", b.attrition_curve_name),
            ):
                got = (raw.get(field_name) or "").strip()
                if got and got != expected:
                    errors.append(
                        ImportError(
                            idx,
                            f"{field_name} differs from first row for trial {name!r} "
                            f"({got!r} vs {expected!r}); leave blank to inherit",
                        )
                    )
            b.assignments.append((site_name, rand, screen))

    # Sum check per trial.
    for b in blocks.values():
        sum_rand = sum(a[1] for a in b.assignments)
        sum_screen = sum(a[2] for a in b.assignments)
        if sum_rand != b.enrollment_target:
            errors.append(
                ImportError(
                    0,
                    f"trial {b.name!r}: per-site rand sum {sum_rand} ≠ "
                    f"study enrollment_target {b.enrollment_target}",
                )
            )
        if sum_screen != b.screening_target:
            errors.append(
                ImportError(
                    0,
                    f"trial {b.name!r}: per-site screen sum {sum_screen} ≠ "
                    f"study screening_target {b.screening_target}",
                )
            )
    return list(blocks.values()), errors


async def preview_trials(text: str, org_id: UUID, db: AsyncSession) -> ImportPreview:
    blocks, errors = await _validate_trials(text, org_id, db)
    actions = []
    for b in blocks:
        actions.append(
            f"Create trial {b.name!r} (draft) with {len(b.assignments)} site assignment(s)"
        )
    return ImportPreview(actions=actions, errors=errors)


async def commit_trials(text: str, org_id: UUID, db: AsyncSession) -> ImportPreview:
    blocks, errors = await _validate_trials(text, org_id, db)
    if errors:
        return ImportPreview(actions=[], errors=errors)

    sites_by_name = {
        s.name: s
        for s in (
            await db.execute(select(Site).where(Site.org_id == org_id))
        ).scalars().all()
    }
    curves_by_name = {
        c.name: c
        for c in (
            await db.execute(
                select(AttritionCurve).where(AttritionCurve.org_id == org_id)
            )
        ).scalars().all()
    }

    actions = []
    for b in blocks:
        trial = Trial(
            org_id=org_id,
            name=b.name,
            sponsor=b.sponsor,
            fpfv=b.fpfv,
            lpfv=b.lpfv,
            lplv=b.lplv,
            enrollment_target=b.enrollment_target,
            screening_target=b.screening_target,
            attrition_curve_id=curves_by_name[b.attrition_curve_name].id,
            status=TrialStatus.DRAFT,
            is_multi_arm=False,
        )
        db.add(trial)
        await db.flush()
        # Single-arm trials get a Default Arm — same pattern as the
        # /trials POST endpoint (see routers/trials.py).
        db.add(Arm(org_id=org_id, trial_id=trial.id, name="Default Arm"))
        for site_name, rand, screen in b.assignments:
            db.add(
                SiteTrial(
                    org_id=org_id,
                    site_id=sites_by_name[site_name].id,
                    trial_id=trial.id,
                    per_site_enrollment_target=rand,
                    per_site_screening_target=screen,
                    active=True,
                )
            )
        actions.append(f"Created trial {b.name!r} with {len(b.assignments)} sites")
    await db.flush()
    return ImportPreview(actions=actions, errors=[])


# ---------------------------------------------------------------------------
# Projections


@dataclass
class _ProjectionRow:
    site_id: UUID
    trial_id: UUID
    arm_id: UUID
    week_start: date
    proj_screened: int
    proj_randomized: int
    site_name: str
    trial_name: str


async def _validate_projections(
    text: str, org_id: UUID, db: AsyncSession
) -> tuple[list[_ProjectionRow], list[ImportError]]:
    errors: list[ImportError] = []
    reader = csv.DictReader(io.StringIO(text))
    required = {
        "site_name", "trial_name", "arm_name", "week_start",
        "proj_screened", "proj_randomized",
    }
    missing = required - set(reader.fieldnames or [])
    if missing:
        errors.append(
            ImportError(1, f"missing required columns: {sorted(missing)}")
        )
        return [], errors

    sites_by_name = {
        s.name: s
        for s in (
            await db.execute(select(Site).where(Site.org_id == org_id))
        ).scalars().all()
    }
    trials_by_name = {
        t.name: t
        for t in (
            await db.execute(select(Trial).where(Trial.org_id == org_id))
        ).scalars().all()
    }
    arms_by_trial_name: dict[UUID, dict[str, Arm]] = {}
    for arm in (await db.execute(select(Arm))).scalars().all():
        arms_by_trial_name.setdefault(arm.trial_id, {})[arm.name] = arm

    out: list[_ProjectionRow] = []
    seen_keys: set[tuple[UUID, UUID, UUID, date]] = set()

    for idx, raw in enumerate(reader, start=2):
        site_name = (raw.get("site_name") or "").strip()
        trial_name = (raw.get("trial_name") or "").strip()
        arm_name = (raw.get("arm_name") or "").strip() or "Default Arm"
        if not site_name or not trial_name:
            errors.append(ImportError(idx, "site_name and trial_name are required"))
            continue
        site = sites_by_name.get(site_name)
        trial = trials_by_name.get(trial_name)
        if site is None:
            errors.append(ImportError(idx, f"unknown site {site_name!r}"))
            continue
        if trial is None:
            errors.append(ImportError(idx, f"unknown trial {trial_name!r}"))
            continue
        arm = arms_by_trial_name.get(trial.id, {}).get(arm_name)
        if arm is None:
            errors.append(
                ImportError(idx, f"unknown arm {arm_name!r} on trial {trial_name!r}")
            )
            continue

        week_start = _parse_date(raw.get("week_start", ""), "week_start", idx, errors)
        if week_start is None:
            continue
        if week_start.weekday() != 0:
            errors.append(
                ImportError(idx, f"week_start {week_start} must be a Monday")
            )
            continue

        proj_s = _parse_int(raw.get("proj_screened", ""), "proj_screened", idx, errors)
        proj_r = _parse_int(
            raw.get("proj_randomized", ""), "proj_randomized", idx, errors
        )
        if proj_s is None or proj_r is None:
            continue
        if proj_s < 0 or proj_r < 0:
            errors.append(ImportError(idx, "projections must be ≥ 0"))
            continue

        key = (site.id, trial.id, arm.id, week_start)
        if key in seen_keys:
            errors.append(
                ImportError(
                    idx,
                    f"duplicate row for ({site_name}, {trial_name}, {arm_name}, "
                    f"{week_start.isoformat()})",
                )
            )
            continue
        seen_keys.add(key)
        out.append(
            _ProjectionRow(
                site_id=site.id,
                trial_id=trial.id,
                arm_id=arm.id,
                week_start=week_start,
                proj_screened=proj_s,
                proj_randomized=proj_r,
                site_name=site_name,
                trial_name=trial_name,
            )
        )
    return out, errors


async def preview_projections(
    text: str, org_id: UUID, db: AsyncSession
) -> ImportPreview:
    rows, errors = await _validate_projections(text, org_id, db)
    actions = [
        f"Upsert projection: {r.site_name} / {r.trial_name} / "
        f"{r.week_start.isoformat()} → screened={r.proj_screened}, "
        f"randomized={r.proj_randomized}"
        for r in rows
    ]
    return ImportPreview(actions=actions, errors=errors)


async def commit_projections(
    text: str, org_id: UUID, db: AsyncSession
) -> ImportPreview:
    rows, errors = await _validate_projections(text, org_id, db)
    if errors:
        return ImportPreview(actions=[], errors=errors)
    # Upsert: if a row exists for the same (site, trial, arm, week), update
    # the projection fields. Actuals are never touched by this import.
    for r in rows:
        existing = (
            await db.execute(
                select(EnrollmentWeek).where(
                    EnrollmentWeek.site_id == r.site_id,
                    EnrollmentWeek.trial_id == r.trial_id,
                    EnrollmentWeek.arm_id == r.arm_id,
                    EnrollmentWeek.week_start == r.week_start,
                )
            )
        ).scalar_one_or_none()
        if existing is None:
            db.add(
                EnrollmentWeek(
                    org_id=org_id,
                    site_id=r.site_id,
                    trial_id=r.trial_id,
                    arm_id=r.arm_id,
                    week_start=r.week_start,
                    proj_screened=r.proj_screened,
                    proj_randomized=r.proj_randomized,
                )
            )
        else:
            existing.proj_screened = r.proj_screened
            existing.proj_randomized = r.proj_randomized
    await db.flush()
    return ImportPreview(
        actions=[f"Wrote {len(rows)} projection week(s)"], errors=[]
    )


# ---------------------------------------------------------------------------
# Dispatch


async def preview(
    kind: ImportKind, text: str, org_id: UUID, db: AsyncSession
) -> ImportPreview:
    if kind == "sites":
        return await preview_sites(text, org_id, db)
    if kind == "trials":
        return await preview_trials(text, org_id, db)
    if kind == "projections":
        return await preview_projections(text, org_id, db)
    raise ValueError(f"unknown import kind: {kind}")


async def commit(
    kind: ImportKind, text: str, org_id: UUID, db: AsyncSession
) -> ImportPreview:
    if kind == "sites":
        return await commit_sites(text, org_id, db)
    if kind == "trials":
        return await commit_trials(text, org_id, db)
    if kind == "projections":
        return await commit_projections(text, org_id, db)
    raise ValueError(f"unknown import kind: {kind}")


# OrgSettings reference kept for completeness — used downstream when we
# add an "import org defaults" kind. Silences linters.
_ = OrgSettings
